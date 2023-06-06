import tkinter as tk
from PIL import Image, ImageTk
import webbrowser
from tkinter import messagebox
import openpyxl


# Create or load the Excel workbook
workbook = openpyxl.Workbook()
try:
    workbook = openpyxl.load_workbook("bmi_records.xlsx")
except FileNotFoundError:
    workbook.save("bmi_records.xlsx")


# Function to open the BMI Category Advice page
def show_category_advice():
    category_advice_page = tk.Toplevel(window)
    category_advice_page.title("Category Advice")

    # Opening text
    opening_text = """
    Welcome to the BMI Category Advice page!
    
    Please select your BMI category to get specific advice and resources.
    """
    opening_label = tk.Label(category_advice_page, text=opening_text, wraplength=300, justify="center")
    opening_label.pack(pady=10)

    # Underweight category button
    underweight_button = tk.Button(category_advice_page, text="Underweight",
                                   command=lambda: webbrowser.open(
                                       "https://www.medicalnewstoday.com/articles/321744#can-diet-help"))
    underweight_button.pack(pady=5)

    # Normal weight category button
    normal_weight_button = tk.Button(category_advice_page, text="Normal Weight",
                                     command=lambda: webbrowser.open(
                                         "https://www.livi.co.uk/your-health/5-simple-tips-for-a-healthy-bmi/"))
    normal_weight_button.pack(pady=5)

    # Overweight category button
    overweight_button = tk.Button(category_advice_page, text="Overweight",
                                  command=lambda: webbrowser.open(
                                      "https://www.everydayhealth.com/diet-nutrition/bmi/how-you-reduce-your-bmi-science-backed-steps/"))
    overweight_button.pack(pady=5)

    # Obesity category button
    obesity_button = tk.Button(category_advice_page, text="Obesity",
                               command=lambda: webbrowser.open("https://www.nhs.uk/conditions/obesity/treatment/"))
    obesity_button.pack(pady=5)

    # Exit button
    exit_button = tk.Button(category_advice_page, text="Exit", command=category_advice_page.destroy)
    exit_button.pack(side="bottom", pady=10)


# Function to display the BMI Calculator page
def show_bmi_calculator():
    bmi_calculator_page = tk.Toplevel(window)
    bmi_calculator_page.title("BMI Calculator")

    # Create input fields
    input_frame = tk.Frame(bmi_calculator_page)
    input_frame.pack()

    name_label = tk.Label(input_frame, text="Name:")
    name_label.grid(row=0, column=0, padx=10, pady=5)
    name_entry = tk.Entry(input_frame)
    name_entry.grid(row=0, column=1, padx=10, pady=5)

    age_label = tk.Label(input_frame, text="Age:")
    age_label.grid(row=1, column=0, padx=10, pady=5)
    age_entry = tk.Entry(input_frame)
    age_entry.grid(row=1, column=1, padx=10, pady=5)

    weight_label = tk.Label(input_frame, text="Weight (kg):")
    weight_label.grid(row=2, column=0, padx=10, pady=5)
    weight_entry = tk.Entry(input_frame)
    weight_entry.grid(row=2, column=1, padx=10, pady=5)

    height_label = tk.Label(input_frame, text="Height (m):")
    height_label.grid(row=3, column=0, padx=10, pady=5)
    height_entry = tk.Entry(input_frame)
    height_entry.grid(row=3, column=1, padx=10, pady=5)

    date_label = tk.Label(input_frame, text="Date:")
    date_label.grid(row=4, column=0, padx=10, pady=5)
    date_entry = tk.Entry(input_frame)
    date_entry.grid(row=4, column=1, padx=10, pady=5)

    # Create the result label
    result_label = tk.Label(bmi_calculator_page, text="", font=("Arial", 12))
    result_label.pack(pady=10)

        # Function to calculate BMI
    def calculate_bmi():
        weight = float(weight_entry.get())
        height = float(height_entry.get())
        bmi = weight / (height ** 2)

        # Categorize BMI
        category = ""
        if bmi < 18.5:
            category = "Underweight"
        elif 18.5 <= bmi < 25:
            category = "Normal Weight"
        elif 25 <= bmi < 30:
            category = "Overweight"
        else:
            category = "Obesity"

        # Save BMI details to Excel
        name = name_entry.get()

        if name in workbook.sheetnames:
            sheet = workbook[name]
            last_row = sheet.max_row + 1
        else:
            sheet = workbook.create_sheet(name)
            sheet["A1"] = "Date"
            sheet["B1"] = "Age"
            sheet["C1"] = "Weight"
            sheet["D1"] = "BMI"
            sheet["E1"] = "BMI Category"
            last_row = 2

        sheet[f"A{last_row}"] = date_entry.get()
        sheet[f"B{last_row}"] = age_entry.get()
        sheet[f"C{last_row}"] = weight_entry.get()
        sheet[f"D{last_row}"] = bmi
        sheet[f"E{last_row}"] = category

        workbook.save("bmi_records.xlsx")

        # Update the result label
        result_text = f"Result:\nBMI: {bmi:.2f}\nCategory: {category}"
        result_label.config(text=result_text)


    # Calculate button
    calculate_button = tk.Button(bmi_calculator_page, text="Calculate BMI", command=calculate_bmi)
    calculate_button.pack(pady=10)

    # Result label
    result_label = tk.Label(bmi_calculator_page, text="", font=("Arial", 12))
    result_label.pack()

    # Exit button
    exit_button = tk.Button(bmi_calculator_page, text="Exit", command=bmi_calculator_page.destroy)
    exit_button.pack(side="bottom", pady=10)


# Create main window
window = tk.Tk()
window.title("BMI Calculator")


# Load and display the opening image
try:
    opening_image = Image.open("BMI_Chart.jpg")
    opening_image = opening_image.resize((700, 400), Image.ANTIALIAS)
    opening_image = ImageTk.PhotoImage(opening_image)
    image_label = tk.Label(window, image=opening_image)
    image_label.pack(pady=10)
except:
    messagebox.showerror("BMI Calculator", "Failed to load the opening image.")



# Opening text
opening_text = """
Welcome to the BMI Calculator!

BMI is used to categorize people’s weight. BMI charts are mainly used for working out the health of populations rather than individuals.

Within a population, there will always be people who are at the extremes (have a high BMI or low BMI).

A high or low BMI may be an indicator of poor diet, varying activity levels, or high stress. Just because someone has a ‘normal BMI’ does not mean that they are healthy.

BMI doesn’t take account of body composition, for example, muscle, fat, bone density. Sex and other factors which can impact your weight can also lead to an inaccurate reading. As such, a BMI calculation is not a suitable measure for some people, including children and young people under 18, pregnant women, and athletes.

Please select an option to proceed.
"""
opening_label = tk.Label(window, text=opening_text, wraplength=700, justify="center")
opening_label.pack()


# Create buttons on the opening page
button_frame = tk.Frame(window)
button_frame.pack()

calculator_button = tk.Button(button_frame, text="BMI Calculator", command=show_bmi_calculator)
calculator_button.grid(row=0, column=0, padx=10, pady=5)

advice_button = tk.Button(button_frame, text="Category Advice", command=show_category_advice)
advice_button.grid(row=0, column=1, padx=10, pady=5)


# Exit button
exit_button = tk.Button(window, text="Exit", command=window.destroy)
exit_button.pack(side="bottom", pady=10)


# Start the main loop
window.mainloop()